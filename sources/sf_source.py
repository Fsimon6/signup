"""顺丰数据来源 - 从客户账单Excel文件获取运费数据"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, Optional

import openpyxl

from utils.data_models import OrderData


class SFSource:
    """顺丰数据来源 - 从客户账单Excel文件读取数据并转换为统一数据模型"""

    def __init__(self):
        self.csv_data = {}
        self.load_excel()

    def load_excel(self):
        """加载客户账单Excel文件"""
        excel_path = Path("excel_source") / "客户账单.csv"
        if not excel_path.exists():
            print(f"  ⚠ 未找到客户账单文件: {excel_path}")
            return

        self.csv_data = {}
        try:
            import zipfile
            with zipfile.ZipFile(excel_path, 'r'):
                pass
            
            temp_path = excel_path.parent / ("temp_" + excel_path.stem + ".xlsx")
            import shutil
            shutil.copy(excel_path, temp_path)
            
            wb = openpyxl.load_workbook(temp_path, data_only=True)
            
            temp_path.unlink()
            ws = wb.active

            headers = {}
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers[header.strip()] = col

            tracking_col = headers.get("顺丰运单号")
            amount_col = headers.get("总费用")

            if not tracking_col or not amount_col:
                print(f"  ✗ 未找到所需列，可用列: {list(headers.keys())}")
                return

            for row in range(2, ws.max_row + 1):
                tracking_number = ws.cell(row=row, column=tracking_col).value
                total_cost = ws.cell(row=row, column=amount_col).value

                if tracking_number is None or total_cost is None:
                    continue

                tracking_number = str(tracking_number).strip()
                total_cost = str(total_cost).strip()

                if not tracking_number or not total_cost:
                    continue

                if tracking_number not in self.csv_data:
                    self.csv_data[tracking_number] = []

                try:
                    self.csv_data[tracking_number].append(float(total_cost))
                except ValueError:
                    continue

            print(f"  ✓ 已加载 {len(self.csv_data)} 个顺丰单号")
        except Exception as exc:
            print(f"  ✗ 加载文件失败: {exc}")

    def query_freight(self, tracking_number: str) -> Optional[OrderData]:
        """查询单个单号的运费

        金额计算逻辑：
        - 如果一个单号对应多条记录（扣款和优惠券返款），将所有金额相加
        - 优惠券返款会以负数形式出现（如-1），会自动抵消
        """
        if tracking_number not in self.csv_data:
            return None

        amounts = self.csv_data[tracking_number]
        total_amount = sum(amounts)

        if len(amounts) > 1:
            print(f"  多条记录: {amounts} → 合计: {total_amount}")

        return OrderData(
            tracking_number=tracking_number,
            freight_amount=str(total_amount),
        )

    def query_freights_batch(self, tracking_numbers: list[str]) -> Dict[str, OrderData]:
        """批量查询多个单号的运费"""
        order_data_map = {}
        for tn in tracking_numbers:
            result = self.query_freight(tn)
            if result:
                order_data_map[tn] = result
        return order_data_map

    def close(self):
        """关闭资源（Excel方式无需关闭）"""
        pass
