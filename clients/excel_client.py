"""Excel文件读取客户端"""

from __future__ import annotations

import openpyxl
from pathlib import Path
from typing import Dict, List, Any


class ExcelClient:
    """Excel文件读取客户端"""

    def __init__(self, file_path: str | Path):
        self.file_path = Path(file_path)
        self.wb = None
        self.ws = None
        self.header_map: Dict[str, int] = {}

    def open(self):
        """打开Excel文件"""
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel文件不存在: {self.file_path}")
        self.wb = openpyxl.load_workbook(self.file_path, data_only=True)
        self.ws = self.wb.active
        self._build_header_map()

    def _build_header_map(self):
        """构建表头到列索引的映射"""
        for col_idx, cell in enumerate(self.ws[1], start=1):
            header = cell.value
            if header:
                self.header_map[str(header).strip()] = col_idx

    def get_column_index(self, column_name: str) -> int | None:
        """获取列索引（从1开始）"""
        return self.header_map.get(column_name)

    def read_column(self, column_name: str, skip_explanation_row: bool = True) -> List[Any]:
        """读取整列数据（不含表头）
        
        Args:
            column_name: 列名
            skip_explanation_row: 是否跳过第2行说明行（默认True）
        """
        col_idx = self.get_column_index(column_name)
        if col_idx is None:
            raise ValueError(f"未找到列: {column_name}")

        start_row = 3 if skip_explanation_row else 2
        values = []
        for row_idx in range(start_row, self.ws.max_row + 1):
            cell = self.ws.cell(row=row_idx, column=col_idx)
            values.append(cell.value)
        return values

    def read_row(self, row_idx: int) -> List[Any]:
        """读取单行数据（从1开始）"""
        return [cell.value for cell in self.ws[row_idx]]

    def get_row_count(self, skip_explanation_row: bool = True) -> int:
        """获取数据行数（不含表头）
        
        Args:
            skip_explanation_row: 是否跳过第2行说明行（默认True）
        """
        skip_rows = 2 if skip_explanation_row else 1
        return max(0, self.ws.max_row - skip_rows)

    def close(self):
        """关闭Excel文件"""
        if self.wb:
            self.wb.close()

    def __enter__(self):
        self.open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    @staticmethod
    def find_latest_excel(dir_path: str | Path) -> Path | None:
        """查找目录中最新的Excel文件"""
        dir_path = Path(dir_path)
        if not dir_path.exists():
            return None

        excel_files = sorted(
            dir_path.glob("*.xlsx"),
            key=lambda f: f.stat().st_mtime,
            reverse=True
        )
        return excel_files[0] if excel_files else None